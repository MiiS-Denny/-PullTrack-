import io
import os
import re
import hmac
from collections import Counter, defaultdict
from datetime import datetime

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from openpyxl.chart import Reference
from hashlib import pbkdf2_hmac

# ---- 時區 ----
try:
    from zoneinfo import ZoneInfo
    TZ = ZoneInfo("Asia/Taipei")
except Exception:
    TZ = None

SHEET_NAME = "Data"
FIXED_BASE = "XbarAndRchart"  # 檔名固定 base

# ---- 雜湊帳密（可改用 st.secrets）----
PWD_DB = {
    "Charles": {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "32ae892164a22af5f83261bd239ed304", "hash": "27fb5fb7bbe2629d8c53dbbdf021423cdb4e7015e5858deafb3a0e405139bb40"},
    "Hsiang":  {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "9c31cdf98b82aa1741154680e456e3e0", "hash": "292e30442d243ea5f82879f1ce71f9ff2dc600f7234a075ba3ee130f45eb29b4"},
    "Sandy":   {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "dd8fbb2a735b076e5cff3bdee67fc3cf", "hash": "7bff0a1388c1447e934552175786d2fa5b9bc9b17ac3d9da246182dd7ec31e35"},
    "Min":     {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "a4a89474d39a1d89ac652a56ccd33301", "hash": "7d788d76be27923209c08aba44fdfc0ca6ce5530ed4b91283810fd0c34bc1a0f"},
    "May":     {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "88d33f6eb3d9a6506b705c3810e7be0b", "hash": "53765f6d56af8c2e49f917c89d60212ab8aeec28d215c9e53cf394e897782631"},
    "Ping":    {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "4af5ee4403ad13cb6a2b0836da5d02b1", "hash": "1c1757b927959d2ef8897467f1c823753ec166f0d5c0a1a8ed5d91a84f2ab00d"},
    "Denny":   {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "bc88ba930b619a25dcce81e6ee616305", "hash": "3dfe81a7dd31acaf2816604c000637f328049d1ca9f13940e217ec51f3a5e7c7"},
    "Davina":  {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "8ce1cb7106316a21db1b48534d7d1833", "hash": "3a79b1feaa96cd7dc7dbced0bc2226d84da22ecda5a38d7d44a58f98e8c24b96"},
    # 如果你已加入 Arthur，保持既有設定
    "Arthur": {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "8e9a0b3e6c6dd1dccd6964101b5af752", "hash": "0409292dedb20de507c7fae67d25f502998c80cb4fcace6758d8fedc042d5570"},

}

def verify_password(username: str, password: str) -> bool:
    rec = PWD_DB.get(username)
    if not rec or rec.get("algo") != "pbkdf2_sha256":
        return False
    salt = bytes.fromhex(rec["salt"])
    digest = pbkdf2_hmac("sha256", password.encode("utf-8"), salt, rec["iter"], dklen=32)
    return hmac.compare_digest(digest, bytes.fromhex(rec["hash"]))

# ---- Excel 工具 ----
COL_DATE = 1; COL_V_START = 2; COL_V_END = 7
COL_XBAR = 8; COL_R = 9
COL_CL_XBAR = 10; COL_UCL_XBAR = 11; COL_LCL_XBAR = 12
COL_CL_R = 13; COL_UCL_R = 14; COL_LCL_R = 15
COL_OWNER = 16

def find_last_data_row(ws, col=COL_DATE):
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=col).value
        if v not in (None, ""):
            return r
    return 1

def validate_yyyymmdd(s: str) -> bool:
    return bool(re.fullmatch(r"\d{8}", s or ""))

def to_float_or_raise(s: str, name: str) -> float:
    try:
        return float(s)
    except Exception:
        raise ValueError(f"{name} 需為數字，收到：{s!r}")

def copy_cell_style(src, dst):
    if src.has_style:
        if src.font:        dst.font = Font(**src.font.__dict__)
        if src.alignment:   dst.alignment = Alignment(**src.alignment.__dict__)
        if src.border:      dst.border = Border(**src.border.__dict__)
        if src.fill:        dst.fill = PatternFill(**src.fill.__dict__)
        if src.protection:  dst.protection = Protection(**src.protection.__dict__)
        dst.number_format = src.number_format

def copy_row_styles(ws, from_row: int, to_row: int, col_start: int, col_end: int):
    ws.row_dimensions[to_row].height = ws.row_dimensions[from_row].height
    for c in range(col_start, col_end + 1):
        copy_cell_style(ws.cell(row=from_row, column=c), ws.cell(row=to_row, column=c))

def _reset_chart_series_to_cols(ws, chart, cols, last_row):
    anchor = getattr(chart, "anchor", None)
    title = getattr(chart, "title", None)
    y_title = getattr(getattr(chart, "y_axis", None), "title", None)
    chart.series = []
    cats = Reference(ws, min_col=COL_DATE, min_row=2, max_row=last_row)
    for col in cols:
        data = Reference(ws, min_col=col, min_row=1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = title
    if hasattr(chart, "y_axis") and y_title is not None:
        chart.y_axis.title = y_title
    if anchor:
        chart.anchor = anchor

def refresh_existing_two_charts(ws, last_row):
    charts = getattr(ws, "_charts", [])
    if not charts: return
    _reset_chart_series_to_cols(ws, charts[0], (COL_XBAR, COL_CL_XBAR, COL_UCL_XBAR, COL_LCL_XBAR), last_row)
    if len(charts) >= 2:
        _reset_chart_series_to_cols(ws, charts[1], (COL_R, COL_CL_R, COL_UCL_R, COL_LCL_R), last_row)

def normalize_value_to_yyyymmdd(v) -> str | None:
    """把任意儲存格值轉成 YYYYMMDD；允許結尾 -N 例如 20251003-2。無法判讀則回 None。"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y%m%d")
    s = str(v).strip()
    # 允許 YYYYMMDD 或 YYYYMMDD-序號
    m = re.fullmatch(r"(\d{8})(?:-(\d+))?", s)
    if m:
        return m.group(1)
    # 也允許 YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD 等
    m = re.fullmatch(r"(\d{4})[-/\.]?(\d{1,2})[-/\.]?(\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}{int(mo):02d}{int(d):02d}"
    return None

def read_last_date_str_from_wb(wb, sheet_name=SHEET_NAME) -> str:
    if sheet_name not in wb.sheetnames: return ""
    ws = wb[sheet_name]
    r = find_last_data_row(ws, col=COL_DATE)
    if r <= 1: return ""
    v = ws.cell(row=r, column=COL_DATE).value
    ymd = normalize_value_to_yyyymmdd(v)
    return ymd or ""

def read_all_dates_from_ws(ws) -> list[str]:
    """讀取 A 欄所有有效日期（第 2 列起），回傳 YYYYMMDD 字串串列（可能重複）。"""
    last = find_last_data_row(ws, col=COL_DATE)
    dates = []
    for r in range(2, last + 1):
        ymd = normalize_value_to_yyyymmdd(ws.cell(row=r, column=COL_DATE).value)
        if ymd:
            dates.append(ymd)
    return dates

def _append_one(ws, date_str: str, values: list[float], owner_text: str, seq_for_same_day: int | None):
    """seq_for_same_day: 若為 1,2,3... 則在 A 欄寫 YYYYMMDD-1 / -2 / ...；None 則寫 YYYYMMDD"""
    last_row = find_last_data_row(ws, col=COL_DATE)
    new_row = last_row + 1
    copy_row_styles(ws, from_row=last_row, to_row=new_row, col_start=COL_DATE, col_end=COL_LCL_R)

    if seq_for_same_day is None:
        ws.cell(row=new_row, column=COL_DATE).value = str(date_str)
    else:
        ws.cell(row=new_row, column=COL_DATE).value = f"{date_str}-{seq_for_same_day}"

    for i, col in enumerate(range(COL_V_START, COL_V_END + 1), start=1):
        ws.cell(row=new_row, column=col).value = values[i-1]

    v_start = ws.cell(row=new_row, column=COL_V_START).coordinate
    v_end   = ws.cell(row=new_row, column=COL_V_END).coordinate
    ws.cell(row=new_row, column=COL_XBAR).value = f"=AVERAGE({v_start}:{v_end})"
    ws.cell(row=new_row, column=COL_R).value    = f"=MAX({v_start}:{v_end})-MIN({v_start}:{v_end})"

    if last_row >= 2:
        for col in (COL_CL_XBAR, COL_UCL_XBAR, COL_LCL_XBAR, COL_CL_R, COL_UCL_R, COL_LCL_R):
            ws.cell(row=new_row, column=col).value = ws.cell(row=last_row, column=col).value

    pcell = ws.cell(row=new_row, column=COL_OWNER)
    pcell.value = owner_text or ""
    pcell.font = Font(name="Calibri", size=11)
    return new_row

def append_many_bytes(template_bytes: bytes, rows_to_add: list, template_name: str, sheet_name=SHEET_NAME,
                      allow_same_day_multi: bool = False, same_day_dates_confirmed: set[str] | None = None):
    """
    回傳：(last_used_date_for_name:str, out_bytes:bytes, reorder_info:dict)
    - 若 allow_same_day_multi=False，遇到與本次輸入彼此或與範本重複的日期會 raise
    - 若 allow_same_day_multi=True，允許同日多筆，並將同日的列在 A 欄標示為 YYYYMMDD-1, -2, ...
      * 僅針對 'same_day_dates_confirmed' 內的日期加 -1/-2，其他仍維持 YYYYMMDD
    """
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"找不到工作表：{sheet_name}")
    ws = wb[sheet_name]

    # 讀取範本資訊
    wb_last_date_str = read_last_date_str_from_wb(wb, sheet_name)
    wb_last_date_int = int(wb_last_date_str) if wb_last_date_str.isdigit() else None
    existing_dates_list = read_all_dates_from_ws(ws)  # A 欄所有已存在日期（YYYYMMDD，可重複）
    existing_dates_set = set(existing_dates_list)

    # 收集本次有填日期的列
    collected = []
    original_order = []
    for idx, row in enumerate(rows_to_add, start=1):
        date_str = (row.get("date") or "").strip()
        if not date_str:
            continue
        if not validate_yyyymmdd(date_str):
            wb.close()
            raise ValueError(f"第 {idx} 列：日期需為 YYYYMMDD，收到 {date_str!r}")
        vals = row.get("values") or []
        if len(vals) != 6:
            wb.close()
            raise ValueError(f"第 {idx} 列：Value_1~Value_6 必須 6 個數字")
        nums = [to_float_or_raise(v, f"第 {idx} 列 Value_{i+1}") for i, v in enumerate(vals)]
        owner = (row.get("owner") or "").strip()
        collected.append({"date": date_str, "values": nums, "owner": owner, "orig_idx": idx})
        original_order.append(date_str)

    if not collected:
        wb.close()
        raise ValueError("沒有可新增的資料：12 列輸入中的日期全為空白")

    # 1) 檢查：本次輸入彼此是否有重複
    cnt = Counter([r["date"] for r in collected])
    dups_input = sorted([d for d, n in cnt.items() if n > 1])

    # 2) 檢查：是否與範本既有日期重複
    dups_with_wb = sorted([d for d in cnt.keys() if d in existing_dates_set])

    # 若不允許同日多筆，且偵測到重複，直接擋下
    need_confirm = bool(dups_input or dups_with_wb)
    if need_confirm and not allow_same_day_multi:
        wb.close()
        # 用統一訊息指出可能的兩種重複來源
        msg_lines = []
        if dups_input:
            msg_lines.append("・本次輸入彼此重複： " + ", ".join(dups_input))
        if dups_with_wb:
            msg_lines.append("・與範本內既有日期重複： " + ", ".join(dups_with_wb))
        raise ValueError("偵測到同日多筆資料。\n" + "\n".join(msg_lines))

    # 自動排序（由小到大）
    sorted_rows = sorted(collected, key=lambda r: int(r["date"]))
    sorted_order = [r["date"] for r in sorted_rows]
    was_reordered = (sorted_order != original_order)

    # 偵測是否有日期早於範本最後一筆日期（僅提示，不擋）
    has_earlier_than_wb = False
    if wb_last_date_int is not None:
        if any(int(r["date"]) < wb_last_date_int for r in sorted_rows):
            has_earlier_than_wb = True

    # 寫入（用排序後的順序）
    last_used_date_for_name = None

    # 決定哪些日期需要加 -1/-2（只對被使用者「確認要多筆」的日期做標註）
    confirmed_set = set(same_day_dates_confirmed or [])

    # 計算本次批次內對於每一個「需標註的日期」的序號
    per_date_running_idx = defaultdict(int)

    for item in sorted_rows:
        d = item["date"]
        seq = None
        if d in confirmed_set:
            per_date_running_idx[d] += 1
            seq = per_date_running_idx[d]  # 1,2,3...
        _append_one(ws, d, item["values"], item["owner"], seq_for_same_day=seq)
        last_used_date_for_name = d

    # 更新圖表資料範圍
    last_row_after = find_last_data_row(ws, col=COL_DATE)
    refresh_existing_two_charts(ws, last_row_after)

    # 輸出 bytes
    out_io = io.BytesIO()
    wb.save(out_io); wb.close()
    out_bytes = out_io.getvalue()

    reorder_info = {
        "original_order": original_order,
        "sorted_order": sorted_order,
        "was_reordered": was_reordered,
        "wb_last_date": wb_last_date_str,
        "has_earlier_than_wb": has_earlier_than_wb,
        "dups_input": dups_input,
        "dups_with_wb": dups_with_wb,
        "confirmed_dates": sorted(list(confirmed_set)),
    }
    return last_used_date_for_name, out_bytes, reorder_info

# ---------------- Streamlit 介面 ----------------
st.set_page_config(page_title="拉力值紀錄（雲端版）", layout="wide")
st.title("拉力值紀錄（雲端版）")

# Session 狀態
if "user" not in st.session_state: st.session_state.user = None
if "seq" not in st.session_state: st.session_state.seq = {}        # prefix -> int
if "last_result" not in st.session_state: st.session_state.last_result = None
if "last_reorder_info" not in st.session_state: st.session_state.last_reorder_info = None
if "login_user" not in st.session_state: st.session_state.login_user = list(PWD_DB.keys())[0]
if "login_pwd" not in st.session_state: st.session_state.login_pwd = ""
if "login_error" not in st.session_state: st.session_state.login_error = ""

# 新增：用於「同日多筆」的互動流程
if "pending_dups" not in st.session_state: st.session_state.pending_dups = None  # dict 存待確認資訊
if "dup_confirmed" not in st.session_state: st.session_state.dup_confirmed = False

# ---- 登入（非 form；Enter 一次就進）----
def attempt_login():
    u = st.session_state.login_user
    p = st.session_state.login_pwd
    if verify_password(u, p):
        st.session_state.user = u
        st.session_state.login_error = ""
        st.rerun()
    else:
        st.session_state.login_error = "帳號或密碼錯誤。"

if st.session_state.user is None:
    st.subheader("登入")
    c1, c2, c3 = st.columns([2, 3, 1.2])
    with c1:
        st.selectbox("帳號", options=list(PWD_DB.keys()), key="login_user")
    with c2:
        st.text_input("密碼", type="password", key="login_pwd", on_change=attempt_login)
    with c3:
        st.write("")
        st.button("登入", on_click=attempt_login, type="primary", use_container_width=True)
    if st.session_state.login_error:
        st.error(st.session_state.login_error)
    st.stop()

# ---- 登出列 ----
logout_col, user_col = st.columns([1, 9])
with logout_col:
    if st.button("登出"):
        st.session_state.user = None
        st.rerun()
with user_col:
    st.write(f"👋 已登入：**{st.session_state.user}**")

st.markdown("---")

# ---- 上傳 Excel ----
st.subheader("① 上傳 Excel 範本（需包含工作表 Data 與兩張圖表）")
tpl_file = st.file_uploader("上傳 .xlsx", type=["xlsx"])

# 預覽最後一筆日期
last_date_placeholder = st.empty()
if tpl_file:
    try:
        wb_preview = load_workbook(tpl_file, data_only=True, read_only=True)
        last_str = read_last_date_str_from_wb(wb_preview, SHEET_NAME)
        wb_preview.close()
        if last_str: last_date_placeholder.info(f"📌 範本目前最後一筆日期：**{last_str}**")
        else:        last_date_placeholder.warning("範本內尚無有效資料或讀不到日期。")
    except Exception as e:
        last_date_placeholder.error(f"讀取範本失敗：{e}")

st.markdown("---")

# ---- 輸入表單（下載鈕在表單外）----
with st.form("input_form"):
    st.subheader("② 輸入資料（一次最多 12 列；空白日期列自動略過）")
    headers = ["Date(YYYYMMDD)", "Value_1(P1-1)", "Value_2(P1-2)", "Value_3(P1-3)",
               "Value_4(P2-1)", "Value_5(P2-2)", "Value_6(P2-3)"]

    rows = []
    for r in range(12):
        c1, c2, c3, c4, c5, c6, c7 = st.columns([1.2, 1, 1, 1, 1, 1, 1])
        with c1: d = st.text_input(f"{headers[0]} #{r+1}", value="", placeholder="YYYYMMDD", key=f"d_{r}")
        with c2: v1 = st.text_input(f"{headers[1]} #{r+1}", value="", key=f"v1_{r}")
        with c3: v2 = st.text_input(f"{headers[2]} #{r+1}", value="", key=f"v2_{r}")
        with c4: v3 = st.text_input(f"{headers[3]} #{r+1}", value="", key=f"v3_{r}")
        with c5: v4 = st.text_input(f"{headers[4]} #{r+1}", value="", key=f"v4_{r}")
        with c6: v5 = st.text_input(f"{headers[5]} #{r+1}", value="", key=f"v5_{r}")
        with c7: v6 = st.text_input(f"{headers[6]} #{r+1}", value="", key=f"v6_{r}")
        rows.append({"date": d.strip(), "values": [v1.strip(), v2.strip(), v3.strip(), v4.strip(), v5.strip(), v6.strip()], "owner": st.session_state.user})

    submitted = st.form_submit_button("③ 送出（生成下載檔）", type="primary", use_container_width=True)

    if submitted:
        if not tpl_file:
            st.error("請先上傳 Excel 範本（.xlsx）。")
        else:
            # 第一次送出：先嘗試不允許同日多筆，若偵測到重複，改用「互動確認」流程
            try:
                # 先只做檢查，不允許同日多筆（遇到重複會 raise）
                _ = append_many_bytes(
                    template_bytes=tpl_file.getvalue(),
                    rows_to_add=rows,
                    template_name=tpl_file.name,
                    sheet_name=SHEET_NAME,
                    allow_same_day_multi=False
                )
                # 若能走到這裡，代表沒有任何重複，直接正式執行（這次才真的產生輸出）
                last_date_added, out_bytes, reorder_info = append_many_bytes(
                    template_bytes=tpl_file.getvalue(),
                    rows_to_add=rows,
                    template_name=tpl_file.name,
                    sheet_name=SHEET_NAME,
                    allow_same_day_multi=True,  # 沒有重複，設 True/False 都可
                    same_day_dates_confirmed=set()
                )

                # 檔名以當下台北時間為準
                now = datetime.now(TZ) if TZ else datetime.now()
                dstr = now.strftime("%Y%m%d")
                hhmm = now.strftime("%H%M")
                prefix = f"{FIXED_BASE}-{dstr}-{hhmm}-"
                n = st.session_state.seq.get(prefix, 0) + 1
                st.session_state.seq[prefix] = n
                out_name = f"{prefix}{n:03d}.xlsx"

                st.session_state.last_result = {
                    "out_name": out_name,
                    "out_bytes": out_bytes,
                    "last_date_added": last_date_added,
                    "generated_at": f"{dstr} {hhmm}",
                }
                st.session_state.last_reorder_info = reorder_info

                st.success(f"已產生：**{out_name}**（請往下滑看下載鈕與提醒）")

            except Exception as e:
                # 如果是「同日多筆」被擋下，進入互動確認流程
                msg = str(e)
                if "偵測到同日多筆資料" in msg:
                    # 解析需要使用者確認的日期清單（可能包含兩種來源）
                    # 從訊息抽出 YYYYMMDD
                    dup_dates = sorted(set(re.findall(r"\b(\d{8})\b", msg)))
                    if not dup_dates:
                        st.error(f"發生錯誤：{e}")
                    else:
                        st.session_state.pending_dups = {
                            "tpl_name": tpl_file.name,
                            "tpl_bytes": tpl_file.getvalue(),
                            "rows": rows,
                            "dup_dates": dup_dates,
                        }
                        st.session_state.last_result = None
                        st.session_state.last_reorder_info = None
                        st.warning("偵測到同日多筆資料，請於下方確認是否允許。")
                else:
                    st.session_state.last_result = None
                    st.session_state.last_reorder_info = None
                    st.error(f"發生錯誤：{e}")

# ---- 「同日多筆」互動區（在表單外）----
if st.session_state.pending_dups:
    st.markdown("---")
    st.subheader("⚠️ 偵測到同日多筆資料")
    pd = st.session_state.pending_dups
    with st.expander("查看重複日期（本次輸入彼此或與範本重複）", expanded=True):
        st.write(", ".join(pd["dup_dates"]))

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("✅ 是，我確認此等日期可有多筆紀錄（將標示 -1/-2/...）", type="primary", use_container_width=True):
            try:
                last_date_added, out_bytes, reorder_info = append_many_bytes(
                    template_bytes=pd["tpl_bytes"],
                    rows_to_add=pd["rows"],
                    template_name=pd["tpl_name"],
                    sheet_name=SHEET_NAME,
                    allow_same_day_multi=True,
                    same_day_dates_confirmed=set(pd["dup_dates"])
                )
                now = datetime.now(TZ) if TZ else datetime.now()
                dstr = now.strftime("%Y%m%d")
                hhmm = now.strftime("%H%M")
                prefix = f"{FIXED_BASE}-{dstr}-{hhmm}-"
                n = st.session_state.seq.get(prefix, 0) + 1
                st.session_state.seq[prefix] = n
                out_name = f"{prefix}{n:03d}.xlsx"

                st.session_state.last_result = {
                    "out_name": out_name,
                    "out_bytes": out_bytes,
                    "last_date_added": last_date_added,
                    "generated_at": f"{dstr} {hhmm}",
                }
                st.session_state.last_reorder_info = reorder_info
                st.session_state.pending_dups = None
                st.success("已允許同日多筆，並完成輸出。請往下滑下載。")
            except Exception as e:
                st.error(f"發生錯誤：{e}")

    with c2:
        if st.button("❌ 否，取消此次輸入", use_container_width=True):
            st.session_state.pending_dups = None
            st.info("已取消此次輸入。")

# ---- 表單外：下載與提醒 ----
if st.session_state.last_result:
    out_name = st.session_state.last_result["out_name"]
    out_bytes = st.session_state.last_result["out_bytes"]
    last_date_added = st.session_state.last_result["last_date_added"]
    generated_at = st.session_state.last_result["generated_at"]

    st.markdown("---")
    st.subheader("③ 下載結果")
    st.success(f"已產生：**{out_name}**")
    st.download_button(
        label="⬇️ 下載檔案",
        data=out_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.info(f"📌 本次實際寫入的最後一筆日期（排序後）：**{last_date_added}**｜產生時間（台北）：**{generated_at}**")

    info = st.session_state.last_reorder_info or {}
    if info.get("was_reordered"):
        st.warning("偵測到輸入日期順序與時間順序不一致，已自動依時間排序再寫入。", icon="⚠️")
        with st.expander("查看原輸入順序 vs 排序後順序"):
            st.write("原輸入順序：", ", ".join(info.get("original_order", [])) or "（無）")
            st.write("排序後順序：", ", ".join(info.get("sorted_order", [])) or "（無）")

    wb_last_date = info.get("wb_last_date")
    if info.get("has_earlier_than_wb"):
        st.warning(
            f"本次輸入中包含早於範本最後日期（{wb_last_date or '未知'}）的紀錄。"
            " 新資料是追加在檔尾，雖已對本次輸入排序，仍可能與檔內既有資料的時序不連續，請留意。",
            icon="🕒"
        )

    confirmed_dates = info.get("confirmed_dates") or []
    if confirmed_dates:
        st.info("以下日期已依序加上 -1 / -2 / ... 標示： " + ", ".join(confirmed_dates))

